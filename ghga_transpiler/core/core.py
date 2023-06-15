# Copyright 2021 - 2023 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

"""This module contains functionalities for processing excel sheets into json object."""
from typing import Union

from openpyxl import Workbook, load_workbook


def read_workbook(filename: str) -> Workbook:
    """
    Function to read-in a workbook (aka spreadsheet)
    """
    return load_workbook(filename)


def get_worksheet_rows(
    worksheet,
    min_row: Optional[int],
    max_row: int,
    min_col: Optional[int],
    max_col: Optional[int],
) -> list:
    """Function to create a list of rows of a worksheet"""
    return list(
        row
        for row in worksheet.iter_rows(
            min_row, max_row, min_col, max_col, values_only=True
        )
        if not all(cell is None for cell in row)
    )


def convert_rows(header: list, rows: list) -> list:
    """Function to return list of dictionaries, rows as values and column names as keys"""
    return [dict(zip(header, row)) for row in rows]


def get_header(worksheet_rows: list[list]) -> list[str]:
    """Function to return a list column names of a worksheet"""
    return worksheet_rows[0]
