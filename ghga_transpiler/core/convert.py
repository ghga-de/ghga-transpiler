# Copyright 2021 - 2023 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""This module contains functionalities for processing excel sheets into json object."""
from openpyxl import load_workbook

from .exceptions import HeaderNotFound


def read_workbook(filename: str):
    """
    Function to read-in spreadsheet
    """
    return load_workbook(filename)


def get_worksheet_rows(
    worksheet, min_row: int, max_row: int, min_col: int, max_col: int
) -> list:
    """Function to generate a list of header values"""
    return list(
        worksheet.iter_rows(min_row, max_row, min_col, max_col, values_only=True)
    )


def _convert_row_to_dict(header: list, row: list) -> dict:
    """
    Function to convert a row into a dictionary with the header as keys.
    """
    return dict(zip(header, row))


def convert_rows(header: list, rows: list) -> list:
    """Function to return list of dictionaries, rows as values and header as keys"""
    return [_convert_row_to_dict(header, row) for row in rows]


def get_sheet_annotation(config, worksheet_name: str) -> dict:
    """Function to return the details of a worksheet structure

    Args:
        config (hexkit...<locals>.ModSettings): config object created from config yaml
    """
    return getattr(config, worksheet_name)


def get_header(sheet_annotation: dict, worksheet_rows: list[list]) -> list[str]:
    """Function to return the header of a worksheet"""
    if sheet_annotation["header"]:
        return worksheet_rows[0]
    raise HeaderNotFound("Worksheet does not have a header")
