# Copyright 2021 - 2023 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

"""This module contains functionalities for processing excel sheets into json object."""
from pathlib import Path
from typing import Tuple, Union

from openpyxl import Workbook, load_workbook

from .config import Config, load_config
from .config.exceptions import MissingWorkbookContent


def read_workbook(filename: str) -> Workbook:
    """
    Function to read-in a workbook (aka spreadsheet)
    """
    return load_workbook(filename)


def get_version(workbook):
    """Function to get workbook version from the worksheet _properties"""
    if "__properties" in workbook.sheetnames:
        return workbook["__properties"].cell(1, 1).value
    print("Using default value: 0.0.1")
    return "0.0.1"


def get_worksheet_rows(
    worksheet,
    min_row: Union[int, None],
    max_row: int,
    min_col: Union[int, None],
    max_col: Union[int, None],
) -> list:
    """Function to create a list of rows of a worksheet"""
    return list(
        row
        for row in worksheet.iter_rows(
            min_row, max_row, min_col, max_col, values_only=True
        )
        if not all(cell is None for cell in row)
    )


def get_header(
    worksheet,
    min_row: Union[int, None],
    min_col: Union[int, None],
    max_col: Union[int, None],
) -> list[str]:
    """Function to return a list column names of a worksheet"""
    return list(
        cell.value
        for row in worksheet.iter_rows(min_row, min_row, min_col, max_col)
        for cell in row
    )


def convert_rows(header, rows: list) -> list:
    """Function to return list of dictionaries, rows as values and column names as keys"""
    return [dict(zip(header, row)) for row in rows]


def params(filename: Path) -> Tuple[Workbook, Config]:
    """Helper function to create workbook, config and channel them into convert_workbook function"""
    workbook = read_workbook(str(filename))
    config = load_config(get_version(workbook))
    return workbook, config


def convert_workbook(workbook: Workbook, config: Config):
    """Function to convert an input spreadsheet into JSON"""
    converted_workbook = {}
    for sheet in config.worksheets:
        if sheet.settings is not None:
            try:
                rows = get_worksheet_rows(
                    workbook[sheet.sheet_name],
                    sheet.settings.start_row,
                    workbook[sheet.sheet_name].max_row,
                    sheet.settings.start_column,
                    sheet.settings.end_column,
                )
            except KeyError as exc:
                raise MissingWorkbookContent(
                    f"Workbook does not contain {sheet.sheet_name} worksheet."
                ) from exc
        else:
            raise ValueError(f"{sheet.settings} cannot be None")
        header = get_header(
            workbook[sheet.sheet_name],
            sheet.settings.header_row,
            sheet.settings.start_column,
            sheet.settings.end_column,
        )
        converted_workbook[sheet.sheet_name] = convert_rows(header, rows)
    return converted_workbook
