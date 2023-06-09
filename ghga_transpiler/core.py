# Copyright 2021 - 2023 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

"""This module contains functionalities for processing excel sheets into json object."""
import re
from importlib import resources
from typing import Callable, Optional, Union

from openpyxl import Workbook

from ghga_transpiler import config

# pylint: disable=line-too-long
SEMVER_REGEX = r"^(0|[1-9]\d*)\.(0|[1-9]\d*)\.(0|[1-9]\d*)(?:-((?:0|[1-9]\d*|\d*[a-zA-Z-][0-9a-zA-Z-]*)(?:\.(?:0|[1-9]\d*|\d*[a-zA-Z-][0-9a-zA-Z-]*))*))?(?:\+([0-9a-zA-Z-]+(?:\.[0-9a-zA-Z-]+)*))?$"


class GHGAWorkbook:
    """A GHGA metadata XLSX workbook"""

    def __init__(self, workbook: Workbook, configs_package: resources.Package):
        """Create a new GHGAWorkbook object from an XLSX workbook"""
        self.workbook = workbook
        self.version = GHGAWorkbook._get_version(workbook)
        self.config = config.load_config(self.version, configs_package)

    @staticmethod
    def _get_version(workbook):
        """Function to get workbook version from the worksheet _properties"""
        if "__properties" in workbook.sheetnames:
            version = str(workbook["__properties"].cell(1, 1).value)
            if re.fullmatch(SEMVER_REGEX, version):
                return version
        raise SyntaxError(
            "Unable to extract metadata version from the provided workbook."
        )


def get_worksheet_rows(
    worksheet,
    min_row: Union[int, None],
    max_row: int,
    min_col: Union[int, None],
    max_col: Union[int, None],
) -> list:
    """Function to create a list of rows of a worksheet"""
    return list(
        row
        for row in worksheet.iter_rows(
            min_row, max_row, min_col, max_col, values_only=True
        )
        if not all(cell is None for cell in row)
    )


def get_header(
    worksheet,
    header_row: Union[int, None],
    min_col: Union[int, None],
    max_col: Union[int, None],
) -> list[str]:
    """Function to return a list column names of a worksheet"""
    return list(
        cell.value
        for row in worksheet.iter_rows(header_row, header_row, min_col, max_col)
        for cell in row
    )


def convert_rows(header, rows) -> list[dict]:
    """Function to return list of dictionaries, rows as worksheet row values and
    column names as keys"""
    return [
        {
            key: value
            for key, value in zip(header, row)
            if value is not None and value != ""
        }
        for row in rows
    ]


def transform_rows(
    rows: list[dict], transformations: Optional[dict[str, Callable]]
) -> list[dict]:
    """Transforms row values if it is applicable with a given function"""
    transformed = []
    for row in rows:
        transformed_row = {}
        for key, value in row.items():
            if transformations and key in transformations:
                transformed_row[key] = transformations[key](value)
            else:
                transformed_row[key] = value
        transformed.append(transformed_row)
    return transformed


def convert_workbook(ghga_workbook: GHGAWorkbook) -> dict:
    """Function to convert an input spreadsheet into JSON"""
    converted_workbook = {}
    for sheet in ghga_workbook.config.worksheets:
        if sheet.settings is not None:
            if sheet.sheet_name in ghga_workbook.workbook:
                rows = get_worksheet_rows(
                    ghga_workbook.workbook[sheet.sheet_name],
                    sheet.settings.start_row,
                    ghga_workbook.workbook[sheet.sheet_name].max_row,
                    sheet.settings.start_column,
                    sheet.settings.end_column,
                )

                header = get_header(
                    ghga_workbook.workbook[sheet.sheet_name],
                    sheet.settings.header_row,
                    sheet.settings.start_column,
                    sheet.settings.end_column,
                )
                converted_rows = convert_rows(header, rows)
                transformed_rows = transform_rows(
                    converted_rows, sheet.settings.transformations
                )
                converted_workbook[sheet.settings.name] = transformed_rows
            else:
                converted_workbook[sheet.settings.name] = []

        else:
            raise ValueError(f"{sheet.settings} will never be None")
    return converted_workbook
