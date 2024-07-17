# Copyright 2021 - 2024 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.


from openpyxl import Workbook
from pydantic import BaseModel

from .config import WorkbookConfig, WorksheetSettings
from .models import GHGAWorkbook, GHGAWorksheet


class WorksheetParser(BaseModel):
    """Group worksheet parser functions."""

    name: str
    config: WorksheetSettings

    def _header(self, worksheet):
        """Return a list of column names of a worksheet."""
        return list(
            cell.value
            for row in worksheet.iter_rows(
                self.config.settings.header_row,
                self.config.settings.header_row,
                self.config.settings.start_column,
                self.config.settings.end_column,
            )
            for cell in row
        )

    def _rows(self, worksheet) -> list:
        """Create a list of rows of a worksheet."""
        return list(
            row
            for row in worksheet.iter_rows(
                self.config.settings.start_row,
                worksheet.max_row,
                self.config.settings.start_column,
                self.config.settings.end_column,
                values_only=True,
            )
            if not all(cell is None for cell in row)
        )

    def _content(self, worksheet) -> list[dict]:
        """Compute and return the content of the worksheet, rows as worksheet row values and
        column names as keys
        """
        return [
            {
                key: value
                for key, value in zip(self._header(worksheet), row, strict=True)
                if value is not None and value != ""
            }
            for row in self._rows(worksheet)
        ]

    def _processed_content(self, worksheet):
        """Transforms row values if it is applicable with a function stated in config"""
        transformed_data = []
        for row in self._content(worksheet):
            transformed_row = {}
            for key, value in row.items():
                transformations = self.config.get_transformations()
                if transformations and key in transformations:
                    transformed_row[key] = transformations[key](value)
                else:
                    transformed_row[key] = value
            transformed_data.append(transformed_row)
        return transformed_data


class GHGAWorksheetParser(WorksheetParser):
    """Extend WorksheetParser with GHGA worksheet specific parsers."""

    def parse(self, worksheet):
        """Function"""
        return GHGAWorksheet.model_validate(
            {"name": self.name, "worksheet": self._parse_worksheet(worksheet)}
        )

    def _parse_worksheet(self, worksheet) -> dict[str, dict]:
        """Parse worksheet row by row into a dictionary of row-primary-keys as keys and
        a dictionary of content and relations as the values.
        """
        worksheet_data = self._processed_content(worksheet)
        return {
            row[self.config.settings.primary_key]: {
                "content": self._relation_free_content(row),
                "relations": self._relations(row),
            }
            for row in worksheet_data
        }

    def _relations(self, row: dict) -> dict:
        """Get relations to a dictionary that contains relation name as key and the
        resource that is in the relation as the value
        """
        relations = self.config.get_relations()
        return {relation: row[relation] for relation in relations if relation in row}

    def _relation_free_content(self, row: dict) -> dict:
        """Clean up the content data from the relation, i.e., remove the key value pairs
        belonging to a relation from the row content.
        """
        return {
            key: value
            for key, value in row.items()
            if key != self.config.settings.primary_key
            and key not in self._relations(row)
        }


class GHGAWorkbookParser:
    """Class"""

    def parse(self, workbook: Workbook, config: WorkbookConfig) -> GHGAWorkbook:
        """Parse a workbook into GHGAWorkbook"""
        return GHGAWorkbook.model_validate(
            {
                "worksheets": tuple(
                    GHGAWorksheetParser(
                        name=name, config=config.worksheets[name]
                    ).parse(workbook[name])
                    for name in workbook.sheetnames
                    if name
                    not in [
                        "__transpiler_protocol",
                        "__sheet_meta",
                        "__column_meta",
                    ]
                )
            }
        )
