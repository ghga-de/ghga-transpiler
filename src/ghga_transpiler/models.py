# Copyright 2021 - 2024 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

"""This module contains functionalities for processing excel sheets into json object."""

import semver
from openpyxl import Workbook
from pydantic import BaseModel

from .config import WorkbookConfig
from .utils import read_meta_information, worksheet_meta_information


class GHGAWorksheetRow(BaseModel):
    """Class"""

    relations: dict
    content: dict


class GHGAWorksheet(BaseModel):
    """Class"""

    name: str
    worksheet: dict[str, GHGAWorksheetRow]


class GHGAWorkbook(BaseModel):
    """Class"""

    worksheets: tuple[GHGAWorksheet, ...]


# class InvalidSematicVersion(Exception):
#     """Raised when a version string is invalid."""


# class GHGAWorkbook:
#     """A GHGA metadata XLSX workbook"""

#     def __init__(self, workbook: Workbook):
#         """Create a new GHGAWorkbook object from an XLSX workbook"""
#         self.workbook = workbook
#         self.wb_version = GHGAWorkbook._get_transpiler_protocol(workbook)
#         self.config = GHGAWorkbook._get_sheet_meta(self.workbook)

#     @staticmethod
#     def _get_transpiler_protocol(workbook):
#         """Gets workbook version from the worksheet "__transpiler_protocol"""
#         if "__transpiler_protocol" in workbook.sheetnames:
#             try:
#                 return semver.Version.parse(
#                     workbook["__transpiler_protocol"].cell(1, 1).value
#                 )
#             except ValueError:
#                 raise InvalidSematicVersion(
#                     "Unable to extract metadata model version from the provided workbook"
#                     "(not a valid semantic version)."
#                 ) from None
#         raise SyntaxError(
#             "Unable to extract metadata model version from the provided workbook (missing)."
#         )

#     @staticmethod
#     def _get_sheet_meta(workbook):
#         """Gets workbook configurations from the worksheet __sheet_meta"""
#         worksheet_meta = worksheet_meta_information(
#             read_meta_information(workbook, "__column_meta"),
#             read_meta_information(workbook, "__sheet_meta"),
#         )
#         return WorkbookConfig.model_validate({"worksheets": worksheet_meta})


# def convert_workbook_to_json(ghga_workbook: GHGAWorkbook) -> dict[str, list[dict]]:
#     """Function to convert an input spreadsheet into JSON"""
#     converted_workbook = {}
#     for name, worksheet in ghga_workbook.config.worksheets.items():
#         if name in ghga_workbook.workbook:
#             rows = get_worksheet_rows(
#                 ghga_workbook.workbook[name],
#                 worksheet.settings.start_row,
#                 ghga_workbook.workbook[name].max_row,
#                 worksheet.settings.start_column,
#                 worksheet.settings.end_column,
#             )

#             header = get_header(
#                 ghga_workbook.workbook[name],
#                 worksheet.settings.header_row,
#                 worksheet.settings.start_column,
#                 worksheet.settings.end_column,
#             )
#             converted_rows = convert_rows(header, rows)
#             transformed_rows = transform_rows(
#                 converted_rows, worksheet.get_transformations()
#             )
#             converted_workbook[name] = transformed_rows
#         else:
#             converted_workbook[name] = []

#     return converted_workbook
